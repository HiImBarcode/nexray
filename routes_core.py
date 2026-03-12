"""
NEXRAY v3 — Core API Routes
All existing endpoints moved from server.py to a FastAPI APIRouter.
"""

from fastapi import APIRouter, Request, HTTPException
from fastapi.responses import JSONResponse
from datetime import datetime, timezone, timedelta
import hashlib
import uuid
import json
import io

from db import (
    get_db, fetchone, fetchall, execute,
    rows_to_list, dict_from_row,
    get_session_user, require_auth, require_role, ROLE_HIERARCHY,
    write_audit,
)

router = APIRouter()


# ===== HEALTH CHECK (unauthenticated) =====
@router.get("/api/health")
async def health_check():
    try:
        with get_db() as db:
            fetchone(db, "SELECT 1 as ok")
        return {"status": "ok", "service": "nexray", "version": "3.0.0"}
    except Exception as e:
        return JSONResponse({"status": "error", "detail": str(e)}, status_code=503)


# ========== AUTH ENDPOINTS ==========

@router.post("/api/auth/login")
async def auth_login(request: Request):
    body = await request.json()
    username = body.get("username", "").strip()
    password = body.get("password", "").strip()
    if not username or not password:
        raise HTTPException(status_code=400, detail="username and password required")

    with get_db() as db:
        user = fetchone(db, "SELECT * FROM users WHERE username=%s AND is_active=1", (username,))
        if not user:
            raise HTTPException(status_code=401, detail="Invalid credentials")
        salt = user["password_salt"] or ""
        expected_hash = hashlib.sha256((salt + password).encode()).hexdigest()
        if user["password_hash"] != expected_hash:
            raise HTTPException(status_code=401, detail="Invalid credentials")

        token = str(uuid.uuid4())
        session_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc)
        expires_at = (now + timedelta(hours=24)).strftime('%Y-%m-%d %H:%M:%S')
        now_str = now.strftime('%Y-%m-%d %H:%M:%S')

        execute(db,
            "INSERT INTO sessions (id, user_id, token, created_at, expires_at) VALUES (%s,%s,%s,%s,%s)",
            (session_id, user["id"], token, now_str, expires_at))
        write_audit(db, user["id"], "login", "user", user["id"])
        db.commit()

        return {
            "token": token,
            "expires_at": expires_at,
            "user": {
                "id": user["id"],
                "username": user["username"],
                "display_name": user["display_name"],
                "email": user["email"],
                "role": user["role"],
                "warehouse_id": user["warehouse_id"],
            }
        }


@router.post("/api/auth/logout")
async def auth_logout(request: Request):
    user = require_auth(request)
    auth = request.headers.get("Authorization", "")
    token = auth[7:] if auth.startswith("Bearer ") else ""
    with get_db() as db:
        execute(db, "DELETE FROM sessions WHERE token=%s", (token,))
        write_audit(db, user["uid"], "logout", "user", user["uid"])
        db.commit()
    return {"success": True}


@router.get("/api/auth/me")
async def auth_me(request: Request):
    user = require_auth(request)
    return {
        "id": user["uid"],
        "username": user["username"],
        "display_name": user["display_name"],
        "email": user["email"],
        "role": user["role"],
        "warehouse_id": user["warehouse_id"],
    }


# ========== DASHBOARD ==========

@router.get("/api/dashboard")
async def get_dashboard(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager', 'warehouse_operator')
    with get_db() as db:
        k = {}
        k['total_active_lots'] = fetchone(db, "SELECT COUNT(*) as c FROM inventory_lots WHERE status='active'")['c']
        k['total_on_hand'] = float(fetchone(db, "SELECT COALESCE(SUM(qty_on_hand),0) as s FROM inventory_lots WHERE status='active'")['s'])
        k['total_reserved'] = float(fetchone(db, "SELECT COALESCE(SUM(qty_reserved),0) as s FROM inventory_lots WHERE status='active'")['s'])
        k['total_available'] = round(k['total_on_hand'] - k['total_reserved'], 2)
        k['low_stock_lots'] = fetchone(db, "SELECT COUNT(*) as c FROM inventory_lots WHERE status='active' AND qty_on_hand < 10")['c']
        k['pending_lines'] = fetchone(db, "SELECT COUNT(*) as c FROM outbound_request_lines WHERE status='pending'")['c']
        k['in_progress_lines'] = fetchone(db, "SELECT COUNT(*) as c FROM outbound_request_lines WHERE status IN ('allocated','in_progress')")['c']
        k['needs_approval'] = fetchone(db, "SELECT COUNT(*) as c FROM outbound_request_lines WHERE status='needs_approval'")['c']
        k['closed_lines'] = fetchone(db, "SELECT COUNT(*) as c FROM outbound_request_lines WHERE status='closed'")['c']
        k['pending_adjustments'] = fetchone(db, "SELECT COUNT(*) as c FROM adjustment_requests WHERE status='pending'")['c']
        k['open_findings'] = fetchone(db, "SELECT COUNT(*) as c FROM reconciliation_findings WHERE resolution_status='open'")['c']
        k['pending_integrations'] = fetchone(db, "SELECT COUNT(*) as c FROM integration_events WHERE status='pending'")['c']
        k['failed_prints'] = fetchone(db, "SELECT COUNT(*) as c FROM print_jobs WHERE status='failed'")['c']

        recent_movements = rows_to_list(fetchall(db, """
            SELECT im.*, il.tracking_id as lot_tracking, i.name as item_name
            FROM inventory_movements im
            LEFT JOIN inventory_lots il ON im.inventory_lot_id = il.id
            LEFT JOIN items i ON il.item_id = i.id ORDER BY im.action_at DESC LIMIT 10
        """))

        recent_findings = rows_to_list(fetchall(db,
            "SELECT * FROM reconciliation_findings ORDER BY created_at DESC LIMIT 10"))

    return {'kpis': k, 'recent_movements': recent_movements, 'recent_findings': recent_findings}


# ========== INVENTORY ==========

@router.get("/api/inventory")
async def get_inventory(request: Request, warehouse_id: str = None, status: str = "active", item_type: str = None):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator', 'manager')
    with get_db() as db:
        query = """SELECT il.*, i.sku, i.name as item_name, i.item_type, i.base_uom,
                   w.code as warehouse_code, w.name as warehouse_name,
                   l.rack_code, l.level_code, l.bin_code, l.location_barcode
                   FROM inventory_lots il
                   LEFT JOIN items i ON il.item_id = i.id LEFT JOIN warehouses w ON il.warehouse_id = w.id
                   LEFT JOIN locations l ON il.location_id = l.id WHERE 1=1"""
        args = []
        if status:
            query += " AND il.status=%s"; args.append(status)
        if warehouse_id:
            query += " AND il.warehouse_id=%s"; args.append(warehouse_id)
        if item_type:
            query += " AND i.item_type=%s"; args.append(item_type)
        query += " ORDER BY il.created_at DESC"
        lots = rows_to_list(fetchall(db, query, args))
        for lot in lots:
            lot['qty_available'] = round((lot['qty_on_hand'] or 0) - (lot['qty_reserved'] or 0), 2)
    return {'lots': lots}


# ========== OUTBOUND ==========

@router.get("/api/outbound")
async def get_outbound(request: Request, status: str = None):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator', 'manager')
    with get_db() as db:
        query = """SELECT orl.*, i.sku, i.name as item_name, orq.reference_no, orq.company_label, orq.warehouse_id, w.code as warehouse_code
                   FROM outbound_request_lines orl LEFT JOIN outbound_requests orq ON orl.outbound_request_id = orq.id
                   LEFT JOIN items i ON orl.item_id = i.id LEFT JOIN warehouses w ON orq.warehouse_id = w.id WHERE 1=1"""
        args = []
        if status:
            query += " AND orl.status=%s"; args.append(status)
        query += " ORDER BY orl.created_at DESC"
        return {'lines': rows_to_list(fetchall(db, query, args))}


@router.get("/api/cuts")
async def get_cuts(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator', 'manager')
    with get_db() as db:
        cuts = rows_to_list(fetchall(db, """
            SELECT ct.*, i.name as item_name, i.sku, il.tracking_id as lot_tracking,
                   orl.qty_requested as line_qty_requested, orl.status as line_status
            FROM cut_transactions ct LEFT JOIN inventory_lots il ON ct.inventory_lot_id = il.id
            LEFT JOIN items i ON il.item_id = i.id LEFT JOIN outbound_request_lines orl ON ct.outbound_request_line_id = orl.id
            ORDER BY ct.cut_at DESC
        """))
    return {'cuts': cuts}


@router.get("/api/tags")
async def get_tags(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator', 'manager')
    with get_db() as db:
        tags = rows_to_list(fetchall(db, """
            SELECT tl.*, ct.qty_actual as cut_qty, il.tracking_id as lot_tracking, i.name as item_name
            FROM tag_labels tl LEFT JOIN cut_transactions ct ON tl.cut_transaction_id = ct.id
            LEFT JOIN inventory_lots il ON tl.inventory_lot_id = il.id LEFT JOIN items i ON il.item_id = i.id
            ORDER BY tl.created_at DESC
        """))
    return {'tags': tags}


@router.get("/api/warehouses")
async def get_warehouses(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator', 'manager')
    with get_db() as db:
        whs = rows_to_list(fetchall(db, """
            SELECT w.*,
                (SELECT COUNT(*) FROM inventory_lots il WHERE il.warehouse_id=w.id AND il.status='active') as active_lots,
                (SELECT COALESCE(SUM(il.qty_on_hand),0) FROM inventory_lots il WHERE il.warehouse_id=w.id AND il.status='active') as total_stock
            FROM warehouses w ORDER BY w.name
        """))
    return {'warehouses': whs}


@router.get("/api/locations")
async def get_locations(request: Request, warehouse_id: str = "wh-01"):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator', 'manager')
    with get_db() as db:
        locs = rows_to_list(fetchall(db, """
            SELECT l.*,
                (SELECT COUNT(*) FROM inventory_lots il WHERE il.location_id=l.id AND il.status='active') as lot_count,
                (SELECT COALESCE(SUM(il.qty_on_hand),0) FROM inventory_lots il WHERE il.location_id=l.id AND il.status='active') as total_qty
            FROM locations l WHERE l.warehouse_id=%s ORDER BY l.zone_code, l.aisle_code, l.rack_code, l.level_code
        """, (warehouse_id,)))
    return {'locations': locs}


@router.get("/api/adjustments")
async def get_adjustments(request: Request, status: str = None):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    with get_db() as db:
        query = "SELECT * FROM adjustment_requests WHERE 1=1"
        args = []
        if status:
            query += " AND status=%s"; args.append(status)
        query += " ORDER BY requested_at DESC"
        return {'adjustments': rows_to_list(fetchall(db, query, args))}


@router.get("/api/findings")
async def get_findings(request: Request, resolution_status: str = None):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    with get_db() as db:
        query = "SELECT * FROM reconciliation_findings WHERE 1=1"
        args = []
        if resolution_status:
            query += " AND resolution_status=%s"; args.append(resolution_status)
        query += " ORDER BY created_at DESC"
        return {'findings': rows_to_list(fetchall(db, query, args))}


@router.get("/api/movements")
async def get_movements(request: Request, lot_id: str = None):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator', 'manager')
    with get_db() as db:
        query = """SELECT im.*, il.tracking_id as lot_tracking, i.name as item_name
                   FROM inventory_movements im LEFT JOIN inventory_lots il ON im.inventory_lot_id = il.id
                   LEFT JOIN items i ON il.item_id = i.id WHERE 1=1"""
        args = []
        if lot_id:
            query += " AND im.inventory_lot_id=%s"; args.append(lot_id)
        query += " ORDER BY im.action_at DESC LIMIT 100"
        return {'movements': rows_to_list(fetchall(db, query, args))}


@router.get("/api/integration_events")
async def get_integration_events(request: Request, status: str = None):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'manager', 'accounting_operator')
    with get_db() as db:
        query = "SELECT * FROM integration_events WHERE 1=1"
        args = []
        if status:
            query += " AND status=%s"; args.append(status)
        query += " ORDER BY created_at DESC"
        return {'events': rows_to_list(fetchall(db, query, args))}


@router.get("/api/users")
async def get_users(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'manager')
    with get_db() as db:
        users = rows_to_list(fetchall(db, """
            SELECT u.id, u.username, u.display_name, u.email, u.role, u.warehouse_id, u.is_active,
                   w.name as warehouse_name
            FROM users u LEFT JOIN warehouses w ON u.warehouse_id = w.id
            ORDER BY u.role, u.display_name
        """))
    return {'users': users}


@router.get("/api/audit_log")
async def get_audit_log(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'manager', 'inventory_admin')
    with get_db() as db:
        logs = rows_to_list(fetchall(db, """
            SELECT al.*, u.display_name as actor_name FROM audit_logs al
            LEFT JOIN users u ON al.actor_user_id = u.id
            ORDER BY al.created_at DESC LIMIT 50
        """))
    return {'logs': logs}


@router.get("/api/supplier_orders")
async def get_supplier_orders(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    with get_db() as db:
        orders = rows_to_list(fetchall(db, """
            SELECT sol.*, s.name as supplier_name FROM supplier_order_lists sol
            LEFT JOIN suppliers s ON sol.supplier_id = s.id ORDER BY sol.created_at DESC
        """))
    return {'orders': orders}


@router.get("/api/supplier_orders/{sol_id}")
async def get_supplier_order(request: Request, sol_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    with get_db() as db:
        order = dict_from_row(fetchone(db, """
            SELECT sol.*, s.name as supplier_name FROM supplier_order_lists sol
            LEFT JOIN suppliers s ON sol.supplier_id = s.id WHERE sol.id=%s
        """, (sol_id,)))
        if not order:
            raise HTTPException(status_code=404, detail="Supplier order not found")
        lines = rows_to_list(fetchall(db,
            "SELECT * FROM supplier_order_list_lines WHERE supplier_order_list_id=%s ORDER BY line_no", (sol_id,)))
        order['lines'] = lines
    return order


@router.get("/api/supplier_orders/{sol_id}/lines")
async def get_supplier_order_lines(request: Request, sol_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator', 'manager')
    with get_db() as db:
        order = dict_from_row(fetchone(db, "SELECT * FROM supplier_order_lists WHERE id=%s", (sol_id,)))
        if not order:
            raise HTTPException(status_code=404, detail="Supplier order not found")
        lines = rows_to_list(fetchall(db,
            "SELECT soll.*, i.sku, i.name as item_name FROM supplier_order_list_lines soll "
            "LEFT JOIN items i ON soll.item_id = i.id WHERE soll.supplier_order_list_id=%s ORDER BY soll.line_no", (sol_id,)))
    return {'order': order, 'lines': lines}


@router.get("/api/print_jobs")
async def get_print_jobs(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator', 'manager')
    with get_db() as db:
        jobs = rows_to_list(fetchall(db, """
            SELECT pj.*, tl.tag_code FROM print_jobs pj
            LEFT JOIN tag_labels tl ON pj.tag_label_id = tl.id ORDER BY pj.created_at DESC
        """))
    return {'jobs': jobs}


@router.get("/api/items")
async def get_items(request: Request, item_type: str = None, is_active: int = 1):
    user = require_auth(request)
    with get_db() as db:
        query = "SELECT * FROM items WHERE is_active=%s"
        args = [is_active]
        if item_type:
            query += " AND item_type=%s"; args.append(item_type)
        query += " ORDER BY name"
        return {'items': rows_to_list(fetchall(db, query, args))}


@router.get("/api/items/{item_id}")
async def get_item(request: Request, item_id: str):
    user = require_auth(request)
    with get_db() as db:
        item = dict_from_row(fetchone(db, "SELECT * FROM items WHERE id=%s", (item_id,)))
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        aliases = rows_to_list(fetchall(db, "SELECT * FROM item_aliases WHERE item_id=%s", (item_id,)))
        item['aliases'] = aliases
    return item


@router.get("/api/suppliers")
async def get_suppliers(request: Request):
    user = require_auth(request)
    with get_db() as db:
        return {'suppliers': rows_to_list(fetchall(db, "SELECT * FROM suppliers ORDER BY name"))}


@router.get("/api/customers")
async def get_customers(request: Request):
    user = require_auth(request)
    with get_db() as db:
        return {'customers': rows_to_list(fetchall(db, "SELECT * FROM customers ORDER BY name"))}


@router.get("/api/uom_conversions")
async def get_uom_conversions(request: Request):
    user = require_auth(request)
    with get_db() as db:
        return {'conversions': rows_to_list(fetchall(db, "SELECT * FROM uom_conversions ORDER BY from_uom, to_uom"))}


# ========== POST ENDPOINTS ==========

@router.post("/api/approve_adjustment")
async def approve_adjustment(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    body = await request.json()
    with get_db() as db:
        before = dict_from_row(fetchone(db, "SELECT * FROM adjustment_requests WHERE id=%s", (body['id'],)))
        execute(db, "UPDATE adjustment_requests SET status='approved', approved_by=%s, approved_at=NOW() WHERE id=%s",
                (user['uid'], body['id']))
        write_audit(db, user['uid'], 'approve', 'adjustment_request', body['id'], before, {'status': 'approved'})
        db.commit()
    return {'success': True}


@router.post("/api/reject_adjustment")
async def reject_adjustment(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    body = await request.json()
    with get_db() as db:
        before = dict_from_row(fetchone(db, "SELECT * FROM adjustment_requests WHERE id=%s", (body['id'],)))
        execute(db, "UPDATE adjustment_requests SET status='rejected', approved_by=%s, approved_at=NOW() WHERE id=%s",
                (user['uid'], body['id']))
        write_audit(db, user['uid'], 'reject', 'adjustment_request', body['id'], before, {'status': 'rejected'})
        db.commit()
    return {'success': True}


@router.post("/api/resolve_finding")
async def resolve_finding(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    body = await request.json()
    with get_db() as db:
        execute(db, "UPDATE reconciliation_findings SET resolution_status='resolved', resolved_by=%s, resolved_at=NOW(), resolution_notes=%s WHERE id=%s",
                (user['uid'], body.get('notes', ''), body['id']))
        db.commit()
    return {'success': True}


@router.post("/api/update_line_status")
async def update_line_status(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator', 'manager')
    body = await request.json()
    with get_db() as db:
        before = dict_from_row(fetchone(db, "SELECT * FROM outbound_request_lines WHERE id=%s", (body['id'],)))
        execute(db, "UPDATE outbound_request_lines SET status=%s, updated_at=NOW() WHERE id=%s",
                (body['status'], body['id']))
        write_audit(db, user['uid'], 'update_status', 'outbound_request_line', body['id'],
                    {'status': before['status'] if before else None}, {'status': body['status']})
        db.commit()
    return {'success': True}


@router.post("/api/retry_integration")
async def retry_integration(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'manager', 'accounting_operator')
    body = await request.json()
    with get_db() as db:
        execute(db, "UPDATE integration_events SET status='pending', retry_count=retry_count+1 WHERE id=%s", (body['id'],))
        db.commit()
    return {'success': True}


# ========== CRUD: WAREHOUSES ==========

@router.post("/api/warehouses")
async def create_warehouse(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin')
    body = await request.json()
    wid = str(uuid.uuid4())
    with get_db() as db:
        execute(db,
            "INSERT INTO warehouses (id, code, name, address, is_active) VALUES (%s,%s,%s,%s,%s)",
            (wid, body['code'], body['name'], body.get('address'), body.get('is_active', 1)))
        write_audit(db, user['uid'], 'create', 'warehouse', wid, None, body)
        db.commit()
    return {'id': wid, 'success': True}


@router.put("/api/warehouses/{warehouse_id}")
async def update_warehouse(request: Request, warehouse_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin')
    body = await request.json()
    with get_db() as db:
        before = dict_from_row(fetchone(db, "SELECT * FROM warehouses WHERE id=%s", (warehouse_id,)))
        if not before:
            raise HTTPException(status_code=404, detail="Warehouse not found")
        fields = {k: v for k, v in body.items() if k in ('code', 'name', 'address', 'is_active')}
        if fields:
            set_clause = ", ".join(f"{k}=%s" for k in fields)
            set_clause += ", updated_at=NOW()"
            execute(db, f"UPDATE warehouses SET {set_clause} WHERE id=%s", list(fields.values()) + [warehouse_id])
        write_audit(db, user['uid'], 'update', 'warehouse', warehouse_id, before, fields)
        db.commit()
    return {'success': True}


# ========== CRUD: LOCATIONS ==========

@router.post("/api/locations")
async def create_location(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin')
    body = await request.json()
    lid = str(uuid.uuid4())
    with get_db() as db:
        execute(db,
            "INSERT INTO locations (id, warehouse_id, zone_code, aisle_code, rack_code, level_code, bin_code, location_barcode, location_type, capacity_qty, is_pickable, is_active) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (lid, body['warehouse_id'], body.get('zone_code'), body.get('aisle_code'), body['rack_code'],
             body.get('level_code'), body.get('bin_code'), body.get('location_barcode'),
             body.get('location_type', 'rack'), body.get('capacity_qty'), body.get('is_pickable', 1), body.get('is_active', 1)))
        write_audit(db, user['uid'], 'create', 'location', lid, None, body)
        db.commit()
    return {'id': lid, 'success': True}


@router.put("/api/locations/{location_id}")
async def update_location(request: Request, location_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin')
    body = await request.json()
    with get_db() as db:
        before = dict_from_row(fetchone(db, "SELECT * FROM locations WHERE id=%s", (location_id,)))
        if not before:
            raise HTTPException(status_code=404, detail="Location not found")
        fields = {k: v for k, v in body.items() if k in ('zone_code', 'aisle_code', 'rack_code', 'level_code', 'bin_code', 'location_barcode', 'location_type', 'capacity_qty', 'is_pickable', 'is_active')}
        if fields:
            set_clause = ", ".join(f"{k}=%s" for k in fields)
            execute(db, f"UPDATE locations SET {set_clause} WHERE id=%s", list(fields.values()) + [location_id])
        write_audit(db, user['uid'], 'update', 'location', location_id, before, fields)
        db.commit()
    return {'success': True}


# ========== CRUD: ITEMS ==========

@router.post("/api/items")
async def create_item(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin')
    body = await request.json()
    iid = str(uuid.uuid4())
    with get_db() as db:
        execute(db,
            "INSERT INTO items (id, sku, name, description, item_type, base_uom, inbound_uom, outbound_uom, category, is_active) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (iid, body['sku'], body['name'], body.get('description'),
             body['item_type'], body.get('base_uom', 'meter'), body.get('inbound_uom', 'meter'),
             body.get('outbound_uom', 'meter'), body.get('category'), body.get('is_active', 1)))
        write_audit(db, user['uid'], 'create', 'item', iid, None, body)
        db.commit()
    return {'id': iid, 'success': True}


@router.put("/api/items/{item_id}")
async def update_item(request: Request, item_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin')
    body = await request.json()
    with get_db() as db:
        before = dict_from_row(fetchone(db, "SELECT * FROM items WHERE id=%s", (item_id,)))
        if not before:
            raise HTTPException(status_code=404, detail="Item not found")
        fields = {k: v for k, v in body.items() if k in ('sku', 'name', 'description', 'item_type', 'base_uom', 'inbound_uom', 'outbound_uom', 'category', 'is_active')}
        if fields:
            set_clause = ", ".join(f"{k}=%s" for k in fields)
            set_clause += ", updated_at=NOW()"
            execute(db, f"UPDATE items SET {set_clause} WHERE id=%s", list(fields.values()) + [item_id])
        write_audit(db, user['uid'], 'update', 'item', item_id, before, fields)
        db.commit()
    return {'success': True}


# ========== CRUD: SUPPLIERS ==========

@router.post("/api/suppliers")
async def create_supplier(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'manager')
    body = await request.json()
    sid = str(uuid.uuid4())
    with get_db() as db:
        execute(db,
            "INSERT INTO suppliers (id, name, code, contact_info, is_active) VALUES (%s,%s,%s,%s,%s)",
            (sid, body['name'], body.get('code'), body.get('contact_info'), body.get('is_active', 1)))
        write_audit(db, user['uid'], 'create', 'supplier', sid, None, body)
        db.commit()
    return {'id': sid, 'success': True}


@router.put("/api/suppliers/{supplier_id}")
async def update_supplier(request: Request, supplier_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'manager')
    body = await request.json()
    with get_db() as db:
        before = dict_from_row(fetchone(db, "SELECT * FROM suppliers WHERE id=%s", (supplier_id,)))
        if not before:
            raise HTTPException(status_code=404, detail="Supplier not found")
        fields = {k: v for k, v in body.items() if k in ('name', 'code', 'contact_info', 'is_active')}
        if fields:
            set_clause = ", ".join(f"{k}=%s" for k in fields)
            execute(db, f"UPDATE suppliers SET {set_clause} WHERE id=%s", list(fields.values()) + [supplier_id])
        write_audit(db, user['uid'], 'update', 'supplier', supplier_id, before, fields)
        db.commit()
    return {'success': True}


# ========== CRUD: CUSTOMERS ==========

@router.post("/api/customers")
async def create_customer(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'manager')
    body = await request.json()
    cid = str(uuid.uuid4())
    with get_db() as db:
        execute(db,
            "INSERT INTO customers (id, name, code, contact_info, is_active) VALUES (%s,%s,%s,%s,%s)",
            (cid, body['name'], body.get('code'), body.get('contact_info'), body.get('is_active', 1)))
        write_audit(db, user['uid'], 'create', 'customer', cid, None, body)
        db.commit()
    return {'id': cid, 'success': True}


@router.put("/api/customers/{customer_id}")
async def update_customer(request: Request, customer_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'manager')
    body = await request.json()
    with get_db() as db:
        before = dict_from_row(fetchone(db, "SELECT * FROM customers WHERE id=%s", (customer_id,)))
        if not before:
            raise HTTPException(status_code=404, detail="Customer not found")
        fields = {k: v for k, v in body.items() if k in ('name', 'code', 'contact_info', 'is_active')}
        if fields:
            set_clause = ", ".join(f"{k}=%s" for k in fields)
            execute(db, f"UPDATE customers SET {set_clause} WHERE id=%s", list(fields.values()) + [customer_id])
        write_audit(db, user['uid'], 'update', 'customer', customer_id, before, fields)
        db.commit()
    return {'success': True}


# ========== CRUD: USERS ==========

@router.post("/api/users")
async def create_user(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin')
    body = await request.json()
    uid = str(uuid.uuid4())
    username = body['username']
    password = body.get('password', username)
    salt = uuid.uuid4().hex[:16]
    phash = hashlib.sha256((salt + password).encode()).hexdigest()
    with get_db() as db:
        execute(db,
            "INSERT INTO users (id, username, display_name, email, password_hash, password_salt, role, warehouse_id, is_active) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (uid, username, body['display_name'], body.get('email'), phash, salt,
             body['role'], body.get('warehouse_id'), body.get('is_active', 1)))
        safe_body = {k: v for k, v in body.items() if k != 'password'}
        write_audit(db, user['uid'], 'create', 'user', uid, None, safe_body)
        db.commit()
    return {'id': uid, 'success': True}


@router.put("/api/users/{user_id}")
async def update_user(request: Request, user_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin')
    body = await request.json()
    with get_db() as db:
        before = dict_from_row(fetchone(db, "SELECT * FROM users WHERE id=%s", (user_id,)))
        if not before:
            raise HTTPException(status_code=404, detail="User not found")
        fields = {k: v for k, v in body.items() if k in ('display_name', 'email', 'role', 'warehouse_id', 'is_active')}
        if fields:
            set_clause = ", ".join(f"{k}=%s" for k in fields)
            set_clause += ", updated_at=NOW()"
            execute(db, f"UPDATE users SET {set_clause} WHERE id=%s", list(fields.values()) + [user_id])
        safe_before = {k: v for k, v in before.items() if k != 'password_hash'}
        write_audit(db, user['uid'], 'update', 'user', user_id, safe_before, fields)
        db.commit()
    return {'success': True}


# ========== INBOUND WORKFLOW ==========

@router.post("/api/supplier_orders")
async def create_supplier_order(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    body = await request.json()
    sol_id = str(uuid.uuid4())
    batch_code = body.get('batch_code') or f"SOL-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"
    lines = body.get('lines', [])
    with get_db() as db:
        execute(db,
            "INSERT INTO supplier_order_lists (id, supplier_id, batch_code, company_label, status, total_lines, notes, created_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
            (sol_id, body.get('supplier_id'), batch_code, body.get('company_label'), 'draft', len(lines), body.get('notes'), user['uid']))
        for idx, line in enumerate(lines, 1):
            line_id = str(uuid.uuid4())
            execute(db,
                "INSERT INTO supplier_order_list_lines (id, supplier_order_list_id, line_no, item_id, item_name_raw, qty_expected, uom, lot_info, shade_info, width_info) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (line_id, sol_id, idx, line.get('item_id'), line.get('item_name_raw'),
                 line['qty_expected'], line.get('uom', 'meter'), line.get('lot_info'),
                 line.get('shade_info'), line.get('width_info')))
        write_audit(db, user['uid'], 'create', 'supplier_order_list', sol_id, None,
                    {'batch_code': batch_code, 'lines': len(lines)})
        db.commit()
    return {'id': sol_id, 'batch_code': batch_code, 'success': True}


@router.put("/api/supplier_orders/{sol_id}")
async def update_supplier_order(request: Request, sol_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    body = await request.json()
    with get_db() as db:
        before = dict_from_row(fetchone(db, "SELECT * FROM supplier_order_lists WHERE id=%s", (sol_id,)))
        if not before:
            raise HTTPException(status_code=404, detail="Supplier order not found")
        fields = {k: v for k, v in body.items() if k in ('supplier_id', 'notes', 'status')}
        if fields:
            set_clause = ", ".join(f"{k}=%s" for k in fields)
            set_clause += ", updated_at=NOW()"
            execute(db, f"UPDATE supplier_order_lists SET {set_clause} WHERE id=%s", list(fields.values()) + [sol_id])
        # Update lines if provided
        new_lines = body.get('lines')
        if new_lines is not None:
            execute(db, "DELETE FROM supplier_order_list_lines WHERE supplier_order_list_id=%s", (sol_id,))
            for idx, line in enumerate(new_lines, 1):
                line_id = str(uuid.uuid4())
                execute(db,
                    "INSERT INTO supplier_order_list_lines (id, supplier_order_list_id, line_no, item_id, item_name_raw, qty_expected, uom, lot_info, shade_info, width_info) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (line_id, sol_id, idx, line.get('item_id'), line.get('item_name_raw'),
                     line['qty_expected'], line.get('uom', 'meter'), line.get('lot_info'),
                     line.get('shade_info'), line.get('width_info')))
            execute(db, "UPDATE supplier_order_lists SET total_lines=%s, updated_at=NOW() WHERE id=%s",
                    (len(new_lines), sol_id))
        write_audit(db, user['uid'], 'update', 'supplier_order_list', sol_id, before, body)
        db.commit()
    return {'success': True}


@router.post("/api/supplier_orders/{sol_id}/validate")
async def validate_supplier_order(request: Request, sol_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead')
    with get_db() as db:
        sol = dict_from_row(fetchone(db, "SELECT * FROM supplier_order_lists WHERE id=%s", (sol_id,)))
        if not sol:
            raise HTTPException(status_code=404, detail="Supplier order not found")
        lines = rows_to_list(fetchall(db, "SELECT * FROM supplier_order_list_lines WHERE supplier_order_list_id=%s", (sol_id,)))
        error_count = 0
        for line in lines:
            err = None
            if not line['item_id']:
                item = fetchone(db, "SELECT id FROM items WHERE name=%s", (line['item_name_raw'],))
                if item:
                    execute(db, "UPDATE supplier_order_list_lines SET item_id=%s WHERE id=%s", (item['id'], line['id']))
                else:
                    err = f"Item not found: '{line['item_name_raw']}'"
            if line['qty_expected'] is None or float(line['qty_expected']) <= 0:
                err = "qty_expected must be > 0"
            if sol.get('supplier_id'):
                sup = fetchone(db, "SELECT id FROM suppliers WHERE id=%s", (sol['supplier_id'],))
                if not sup:
                    err = f"Supplier {sol['supplier_id']} not found"
            if err:
                error_count += 1
                execute(db, "UPDATE supplier_order_list_lines SET validation_status='error', validation_error=%s WHERE id=%s",
                        (err, line['id']))
            else:
                execute(db, "UPDATE supplier_order_list_lines SET validation_status='valid', validation_error=NULL WHERE id=%s",
                        (line['id'],))
        new_status = 'validated' if error_count == 0 else 'failed_with_errors'
        execute(db, "UPDATE supplier_order_lists SET status=%s, error_count=%s, updated_at=NOW() WHERE id=%s",
                (new_status, error_count, sol_id))
        write_audit(db, user['uid'], 'validate', 'supplier_order_list', sol_id,
                    {'status': sol['status']}, {'status': new_status, 'error_count': error_count})
        db.commit()
    return {'success': True, 'status': new_status, 'error_count': error_count}


@router.post("/api/supplier_orders/import")
async def import_supplier_order(request: Request):
    """Import supplier order from CSV or XLSX file."""
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')

    import csv

    form = await request.form()
    file = form.get('file')
    supplier_id = form.get('supplier_id')
    notes = form.get('notes', '')

    if not file:
        raise HTTPException(status_code=400, detail="File is required")

    filename = file.filename.lower()
    content = await file.read()
    parsed_lines = []

    if filename.endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            parsed_lines.append({
                'item_name_raw': row.get('item_name') or row.get('item_name_raw') or row.get('sku') or '',
                'qty_expected': float(row.get('qty_expected') or row.get('qty') or 0),
                'uom': row.get('uom') or 'meter',
                'shade_info': row.get('shade') or row.get('shade_info') or None,
                'width_info': row.get('width') or row.get('width_info') or None,
            })
    elif filename.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(cell.value or '').strip().lower() for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                parsed_lines.append({
                    'item_name_raw': str(row_dict.get('item_name') or row_dict.get('item_name_raw') or row_dict.get('sku') or ''),
                    'qty_expected': float(row_dict.get('qty_expected') or row_dict.get('qty') or 0),
                    'uom': str(row_dict.get('uom') or 'meter'),
                    'shade_info': str(row_dict.get('shade') or row_dict.get('shade_info') or '') or None,
                    'width_info': str(row_dict.get('width') or row_dict.get('width_info') or '') or None,
                })
            wb.close()
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed for XLSX parsing")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Use .csv or .xlsx")

    if not parsed_lines:
        raise HTTPException(status_code=400, detail="No data rows found in file")

    sol_id = str(uuid.uuid4())
    batch_code = f"IMP-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"

    with get_db() as db:
        execute(db,
            "INSERT INTO supplier_order_lists (id, supplier_id, batch_code, status, total_lines, notes, created_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (sol_id, supplier_id, batch_code, 'draft', len(parsed_lines), notes, user['uid']))

        # Resolve items by name/sku
        for idx, line in enumerate(parsed_lines, 1):
            line_id = str(uuid.uuid4())
            item_id = None
            item_name_raw = line['item_name_raw']
            if item_name_raw:
                item = fetchone(db, "SELECT id FROM items WHERE sku=%s OR name=%s", (item_name_raw, item_name_raw))
                if item:
                    item_id = item['id']
            execute(db,
                "INSERT INTO supplier_order_list_lines (id, supplier_order_list_id, line_no, item_id, item_name_raw, qty_expected, uom, shade_info, width_info) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (line_id, sol_id, idx, item_id, item_name_raw,
                 line['qty_expected'], line['uom'], line.get('shade_info'), line.get('width_info')))

        write_audit(db, user['uid'], 'import', 'supplier_order_list', sol_id, None,
                    {'batch_code': batch_code, 'lines': len(parsed_lines), 'filename': file.filename})
        db.commit()

    return {'id': sol_id, 'batch_code': batch_code, 'lines_imported': len(parsed_lines), 'success': True}


@router.post("/api/receivings")
async def create_receiving(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator')
    body = await request.json()
    rid = str(uuid.uuid4())
    with get_db() as db:
        execute(db,
            "INSERT INTO receivings (id, warehouse_id, supplier_order_list_id, status, received_by, notes) "
            "VALUES (%s,%s,%s,%s,%s,%s)",
            (rid, body['warehouse_id'], body.get('supplier_order_list_id'), 'in_progress', user['uid'], body.get('notes')))
        if body.get('supplier_order_list_id'):
            execute(db, "UPDATE supplier_order_lists SET status='receiving', updated_at=NOW() WHERE id=%s",
                    (body['supplier_order_list_id'],))

        # If lines provided (Odoo-style), create lots for each confirmed line
        recv_lines = body.get('lines', [])
        created_lots = []
        for line in recv_lines:
            lot_id = str(uuid.uuid4())
            tracking_id = line.get('tracking_id') or f"TRK-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}-{lot_id[:6].upper()}"
            qty = float(line['qty_original'])
            execute(db,
                "INSERT INTO inventory_lots (id, item_id, tracking_id, batch_no, shade_code, width_value, "
                "qty_original, qty_on_hand, qty_reserved, warehouse_id, location_id, status, qty_confidence, receiving_id, supplier_order_line_id, created_by) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (lot_id, line['item_id'], tracking_id, line.get('batch_no'), line.get('shade_code'), line.get('width_value'),
                 qty, qty, 0.0, body['warehouse_id'], line.get('location_id'),
                 'active', line.get('qty_confidence', 'supplier_reported'),
                 rid, line.get('supplier_order_line_id'), user['uid']))
            mov_id = str(uuid.uuid4())
            execute(db,
                "INSERT INTO inventory_movements (id, event_idempotency_key, movement_type, inventory_lot_id, "
                "tracking_id, qty_delta, qty_before, qty_after, warehouse_to_id, location_to_id, action_by) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (mov_id, f"recv-{lot_id}", 'receive', lot_id, tracking_id,
                 qty, 0.0, qty, body['warehouse_id'], line.get('location_id'), user['uid']))
            created_lots.append({'lot_id': lot_id, 'tracking_id': tracking_id})

        write_audit(db, user['uid'], 'create', 'receiving', rid, None, body)
        db.commit()
    return {'id': rid, 'created_lots': created_lots, 'success': True}


@router.post("/api/receivings/{receiving_id}/receive_lot")
async def receive_lot(request: Request, receiving_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator')
    body = await request.json()
    with get_db() as db:
        recv = dict_from_row(fetchone(db, "SELECT * FROM receivings WHERE id=%s", (receiving_id,)))
        if not recv:
            raise HTTPException(status_code=404, detail="Receiving session not found")
        if recv['status'] != 'in_progress':
            raise HTTPException(status_code=400, detail="Receiving session is not in progress")

        lot_id = str(uuid.uuid4())
        tracking_id = body.get('tracking_id') or f"TRK-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}-{lot_id[:6].upper()}"
        qty = float(body['qty_original'])

        execute(db,
            "INSERT INTO inventory_lots (id, item_id, tracking_id, batch_no, shade_code, width_value, "
            "qty_original, qty_on_hand, qty_reserved, warehouse_id, location_id, status, qty_confidence, receiving_id, supplier_order_line_id, created_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (lot_id, body['item_id'], tracking_id, body.get('batch_no'),
             body.get('shade_code'), body.get('width_value'),
             qty, qty, 0.0, recv['warehouse_id'], body.get('location_id'),
             'active', body.get('qty_confidence', 'supplier_reported'),
             receiving_id, body.get('supplier_order_line_id'), user['uid']))

        mov_id = str(uuid.uuid4())
        ikey = f"recv-{lot_id}"
        execute(db,
            "INSERT INTO inventory_movements (id, event_idempotency_key, movement_type, inventory_lot_id, "
            "tracking_id, qty_delta, qty_before, qty_after, warehouse_to_id, location_to_id, action_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (mov_id, ikey, 'receive', lot_id, tracking_id,
             qty, 0.0, qty, recv['warehouse_id'], body.get('location_id'), user['uid']))

        if body.get('supplier_order_line_id'):
            execute(db, "UPDATE supplier_order_list_lines SET validation_status='valid' WHERE id=%s",
                    (body['supplier_order_line_id'],))

        write_audit(db, user['uid'], 'receive_lot', 'inventory_lot', lot_id,
                    None, {'tracking_id': tracking_id, 'qty': qty})
        db.commit()
    return {'lot_id': lot_id, 'tracking_id': tracking_id, 'success': True}


@router.post("/api/receivings/{receiving_id}/confirm")
async def confirm_receiving(request: Request, receiving_id: str):
    """Odoo-style confirm: finalize all lots in receiving session at once."""
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead')
    with get_db() as db:
        recv = dict_from_row(fetchone(db, "SELECT * FROM receivings WHERE id=%s", (receiving_id,)))
        if not recv:
            raise HTTPException(status_code=404, detail="Receiving session not found")
        execute(db, "UPDATE receivings SET status='completed' WHERE id=%s", (receiving_id,))
        if recv.get('supplier_order_list_id'):
            execute(db, "UPDATE supplier_order_lists SET status='completed', updated_at=NOW() WHERE id=%s",
                    (recv['supplier_order_list_id'],))
        write_audit(db, user['uid'], 'confirm', 'receiving', receiving_id,
                    {'status': 'in_progress'}, {'status': 'completed'})
        db.commit()
    return {'success': True}


@router.post("/api/receivings/{receiving_id}/complete")
async def complete_receiving(request: Request, receiving_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead')
    with get_db() as db:
        recv = dict_from_row(fetchone(db, "SELECT * FROM receivings WHERE id=%s", (receiving_id,)))
        if not recv:
            raise HTTPException(status_code=404, detail="Receiving session not found")
        execute(db, "UPDATE receivings SET status='completed' WHERE id=%s", (receiving_id,))
        if recv.get('supplier_order_list_id'):
            execute(db, "UPDATE supplier_order_lists SET status='completed', updated_at=NOW() WHERE id=%s",
                    (recv['supplier_order_list_id'],))
        write_audit(db, user['uid'], 'complete', 'receiving', receiving_id,
                    {'status': 'in_progress'}, {'status': 'completed'})
        db.commit()
    return {'success': True}


@router.post("/api/putaway")
async def putaway_lot(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator')
    body = await request.json()
    tracking_id = body.get('tracking_id')
    location_id = body.get('location_id')
    if not tracking_id or not location_id:
        raise HTTPException(status_code=400, detail="tracking_id and location_id required")
    with get_db() as db:
        lot = dict_from_row(fetchone(db, "SELECT * FROM inventory_lots WHERE tracking_id=%s", (tracking_id,)))
        if not lot:
            raise HTTPException(status_code=404, detail="Lot not found")
        loc = dict_from_row(fetchone(db, "SELECT * FROM locations WHERE id=%s", (location_id,)))
        if not loc:
            raise HTTPException(status_code=404, detail="Location not found")

        from_location_id = lot['location_id']
        pe_id = str(uuid.uuid4())
        execute(db,
            "INSERT INTO putaway_events (id, inventory_lot_id, from_location_id, to_location_id, qty_moved, moved_by, status) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (pe_id, lot['id'], from_location_id, location_id, lot['qty_on_hand'], user['uid'], 'completed'))
        execute(db, "UPDATE inventory_lots SET location_id=%s, updated_at=NOW() WHERE id=%s",
                (location_id, lot['id']))

        mov_id = str(uuid.uuid4())
        ikey = f"putaway-{pe_id}"
        execute(db,
            "INSERT INTO inventory_movements (id, event_idempotency_key, movement_type, inventory_lot_id, "
            "tracking_id, qty_delta, qty_before, qty_after, warehouse_from_id, location_from_id, warehouse_to_id, location_to_id, action_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (mov_id, ikey, 'move', lot['id'], tracking_id,
             0.0, lot['qty_on_hand'], lot['qty_on_hand'],
             lot['warehouse_id'], from_location_id, lot['warehouse_id'], location_id, user['uid']))
        write_audit(db, user['uid'], 'putaway', 'inventory_lot', lot['id'],
                    {'location_id': from_location_id}, {'location_id': location_id})
        db.commit()
    return {'putaway_event_id': pe_id, 'success': True}


# ========== OUTBOUND WORKFLOW ==========

@router.post("/api/outbound_requests")
async def create_outbound_request(request: Request):
    """Create outbound request with lines directly (no batch wrapper)."""
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    body = await request.json()
    or_id = str(uuid.uuid4())
    lines = body.get('lines', [])
    with get_db() as db:
        execute(db,
            "INSERT INTO outbound_requests (id, warehouse_id, customer_id, company_label, reference_no, status, created_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (or_id, body['warehouse_id'], body.get('customer_id'), body.get('company_label'), body.get('reference_no'), 'pending', user['uid']))
        for idx, line in enumerate(lines, 1):
            line_id = str(uuid.uuid4())
            execute(db,
                "INSERT INTO outbound_request_lines (id, outbound_request_id, line_no, item_id, item_name_raw, qty_requested, uom, status) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                (line_id, or_id, idx, line.get('item_id'), line.get('item_name_raw'),
                 line['qty'], line.get('uom', 'meter'), 'pending'))
        write_audit(db, user['uid'], 'create', 'outbound_request', or_id, None,
                    {'reference_no': body.get('reference_no'), 'total_lines': len(lines)})
        db.commit()
    return {'id': or_id, 'total_lines': len(lines), 'success': True}


@router.post("/api/outbound/allocate")
async def allocate_outbound_line(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead')
    body = await request.json()
    line_id = body['line_id']
    with get_db() as db:
        line = dict_from_row(fetchone(db,
            "SELECT orl.*, orq.warehouse_id "
            "FROM outbound_request_lines orl JOIN outbound_requests orq ON orl.outbound_request_id = orq.id "
            "WHERE orl.id=%s", (line_id,)))
        if not line:
            raise HTTPException(status_code=404, detail="Line not found")
        qty_needed = float(line['qty_requested']) - float(line.get('qty_allocated') or 0)
        if qty_needed <= 0:
            return {'success': True, 'message': 'Already fully allocated'}

        lots = rows_to_list(fetchall(db,
            "SELECT * FROM inventory_lots WHERE item_id=%s AND warehouse_id=%s AND status='active' "
            "AND (qty_on_hand - qty_reserved) > 0 ORDER BY created_at ASC",
            (line['item_id'], line['warehouse_id'])))

        total_allocated = 0.0
        reservations_made = []
        for lot in lots:
            if qty_needed <= 0:
                break
            available = float(lot['qty_on_hand']) - float(lot.get('qty_reserved') or 0)
            if available <= 0:
                continue
            reserve_qty = min(available, qty_needed)
            res_id = str(uuid.uuid4())
            execute(db,
                "INSERT INTO inventory_reservations (id, inventory_lot_id, outbound_request_line_id, qty_reserved, status, created_by) "
                "VALUES (%s,%s,%s,%s,%s,%s)",
                (res_id, lot['id'], line_id, reserve_qty, 'active', user['uid']))
            execute(db, "UPDATE inventory_lots SET qty_reserved = qty_reserved + %s, updated_at=NOW() WHERE id=%s",
                    (reserve_qty, lot['id']))
            total_allocated += reserve_qty
            qty_needed -= reserve_qty
            reservations_made.append({'lot_id': lot['id'], 'qty': reserve_qty})

        execute(db,
            "UPDATE outbound_request_lines SET qty_allocated = qty_allocated + %s, status='allocated', updated_at=NOW() WHERE id=%s",
            (total_allocated, line_id))
        write_audit(db, user['uid'], 'allocate', 'outbound_request_line', line_id,
                    {'qty_allocated': line.get('qty_allocated', 0)},
                    {'qty_allocated': float(line.get('qty_allocated', 0)) + total_allocated, 'reservations': reservations_made})
        db.commit()
    return {'success': True, 'total_allocated': total_allocated, 'reservations': reservations_made}


@router.post("/api/outbound/claim_line")
async def claim_outbound_line(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator')
    body = await request.json()
    line_id = body['line_id']
    with get_db() as db:
        line = dict_from_row(fetchone(db, "SELECT * FROM outbound_request_lines WHERE id=%s", (line_id,)))
        if not line:
            raise HTTPException(status_code=404, detail="Line not found")
        lock = fetchone(db,
            "SELECT * FROM workflow_locks WHERE resource_type='outbound_request_line' AND resource_id=%s", (line_id,))
        if lock:
            raise HTTPException(status_code=409, detail=f"Line already claimed by {lock['lock_owner']}")

        now = datetime.now(timezone.utc)
        expires_at = (now + timedelta(hours=2)).strftime('%Y-%m-%d %H:%M:%S')
        lock_id = str(uuid.uuid4())
        execute(db,
            "INSERT INTO workflow_locks (id, resource_type, resource_id, lock_owner, lock_expires_at) VALUES (%s,%s,%s,%s,%s)",
            (lock_id, 'outbound_request_line', line_id, user['uid'], expires_at))
        execute(db,
            "UPDATE outbound_request_lines SET claimed_by=%s, claimed_at=NOW(), status='in_progress', updated_at=NOW() WHERE id=%s",
            (user['uid'], line_id))
        write_audit(db, user['uid'], 'claim', 'outbound_request_line', line_id,
                    {'status': line['status']}, {'status': 'in_progress', 'claimed_by': user['uid']})
        db.commit()
    return {'success': True, 'lock_id': lock_id, 'expires_at': expires_at}


@router.post("/api/outbound/record_cut")
async def record_cut(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator')
    body = await request.json()
    line_id = body['line_id']
    lot_id = body['lot_id']
    qty_requested = float(body['qty_requested'])
    qty_actual = float(body['qty_actual'])
    cut_by = body.get('cut_by', user['uid'])
    qty_variance = round(qty_actual - qty_requested, 4)

    with get_db() as db:
        line = dict_from_row(fetchone(db, "SELECT * FROM outbound_request_lines WHERE id=%s", (line_id,)))
        if not line:
            raise HTTPException(status_code=404, detail="Line not found")
        lot = dict_from_row(fetchone(db, "SELECT * FROM inventory_lots WHERE id=%s", (lot_id,)))
        if not lot:
            raise HTTPException(status_code=404, detail="Lot not found")

        cut_id = str(uuid.uuid4())
        execute(db,
            "INSERT INTO cut_transactions (id, outbound_request_line_id, inventory_lot_id, tracking_id, "
            "qty_requested, qty_actual, qty_variance, variance_reason, cut_by, status) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (cut_id, line_id, lot_id, lot['tracking_id'],
             qty_requested, qty_actual, qty_variance, body.get('variance_reason'), cut_by, 'recorded'))

        # Deduct inventory
        qty_before = float(lot['qty_on_hand'])
        qty_after = round(qty_before - qty_actual, 4)
        mov_id = str(uuid.uuid4())
        ikey = f"cut-{cut_id}-deduct"
        execute(db,
            "INSERT INTO inventory_movements (id, event_idempotency_key, movement_type, inventory_lot_id, "
            "tracking_id, sales_order_line_id, cut_transaction_id, qty_delta, qty_before, qty_after, "
            "warehouse_from_id, location_from_id, action_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (mov_id, ikey, 'deduct', lot_id, lot['tracking_id'],
             line_id, cut_id, -qty_actual, qty_before, qty_after,
             lot['warehouse_id'], lot.get('location_id'), cut_by))
        execute(db, "UPDATE inventory_lots SET qty_on_hand=%s, updated_at=NOW() WHERE id=%s",
                (qty_after, lot_id))

        # Auto-generate tag
        tag_id = str(uuid.uuid4())
        tag_count_row = fetchone(db, "SELECT COUNT(*) as c FROM tag_labels")
        tag_count = tag_count_row['c'] if tag_count_row else 0
        tag_code = f"NXR-TAG-{(tag_count + 1):04d}"
        execute(db,
            "INSERT INTO tag_labels (id, tag_code, cut_transaction_id, inventory_lot_id, outbound_request_line_id, tag_status) "
            "VALUES (%s,%s,%s,%s,%s,%s)",
            (tag_id, tag_code, cut_id, lot_id, line_id, 'generated'))
        pj_id = str(uuid.uuid4())
        execute(db,
            "INSERT INTO print_jobs (id, tag_label_id, job_type, status) VALUES (%s,%s,%s,%s)",
            (pj_id, tag_id, 'tag', 'queued'))

        # Multi-cut: track total fulfilled as sum of all cuts for this line
        total_fulfilled_row = fetchone(db,
            "SELECT COALESCE(SUM(qty_actual),0) as total FROM cut_transactions WHERE outbound_request_line_id=%s AND status != 'voided'",
            (line_id,))
        total_fulfilled = float(total_fulfilled_row['total']) if total_fulfilled_row else qty_actual

        # Check variance: if abs variance > 5%, flag for approval
        variance_pct = abs(qty_actual - qty_requested) / qty_requested if qty_requested > 0 else 0
        line_needs_approval = variance_pct > 0.05
        if line_needs_approval:
            adj_id = str(uuid.uuid4())
            execute(db,
                "INSERT INTO adjustment_requests (id, inventory_lot_id, outbound_request_line_id, "
                "adjustment_type, qty_before, qty_after, reason_code, notes, status, requested_by) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (adj_id, lot_id, line_id, 'variance_approval',
                 qty_requested, qty_actual, 'cut_variance',
                 f"Cut variance {variance_pct*100:.1f}% exceeds 5% threshold", 'pending', cut_by))
            execute(db, "UPDATE outbound_request_lines SET qty_fulfilled=%s, status='needs_approval', updated_at=NOW() WHERE id=%s",
                    (total_fulfilled, line_id))
        else:
            # Only set cut_complete when total fulfilled >= qty_requested
            new_status = 'cut_complete' if total_fulfilled >= float(line['qty_requested']) else 'in_progress'
            execute(db, "UPDATE outbound_request_lines SET qty_fulfilled=%s, status=%s, updated_at=NOW() WHERE id=%s",
                    (total_fulfilled, new_status, line_id))

        write_audit(db, user['uid'], 'record_cut', 'cut_transaction', cut_id,
                    None, {'qty_requested': qty_requested, 'qty_actual': qty_actual, 'tag_code': tag_code})
        db.commit()
    return {
        'cut_id': cut_id, 'tag_id': tag_id, 'tag_code': tag_code, 'print_job_id': pj_id,
        'needs_approval': line_needs_approval, 'total_fulfilled': total_fulfilled, 'success': True
    }


@router.post("/api/outbound/close_line")
async def close_line(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    body = await request.json()
    line_id = body['line_id']
    with get_db() as db:
        line = dict_from_row(fetchone(db, "SELECT * FROM outbound_request_lines WHERE id=%s", (line_id,)))
        if not line:
            raise HTTPException(status_code=404, detail="Line not found")
        # GATE CHECK: at least one cut must exist
        cuts = rows_to_list(fetchall(db,
            "SELECT id FROM cut_transactions WHERE outbound_request_line_id=%s AND status != 'voided'", (line_id,)))
        if not cuts:
            raise HTTPException(status_code=400, detail="Gate check failed: no cuts recorded for this line")

        # Release reservations
        execute(db,
            "UPDATE inventory_reservations SET status='released' WHERE outbound_request_line_id=%s AND status='active'",
            (line_id,))
        execute(db, "DELETE FROM workflow_locks WHERE resource_type='outbound_request_line' AND resource_id=%s", (line_id,))

        execute(db,
            "UPDATE outbound_request_lines SET status='closed', fulfilled_by=%s, fulfilled_at=NOW(), updated_at=NOW() WHERE id=%s",
            (user['uid'], line_id))

        ie_id = str(uuid.uuid4())
        ikey = f"close-line-{line_id}"
        payload = json.dumps({'line_id': line_id, 'qty_fulfilled': line.get('qty_fulfilled', 0)})
        try:
            execute(db,
                "INSERT INTO integration_events (id, event_type, event_idempotency_key, payload_json, status, direction) "
                "VALUES (%s,%s,%s,%s,%s,%s)",
                (ie_id, 'fulfillment_complete', ikey, payload, 'pending', 'outbound'))
        except Exception:
            pass

        write_audit(db, user['uid'], 'close', 'outbound_request_line', line_id,
                    {'status': line['status']}, {'status': 'closed'})
        db.commit()
    return {'success': True, 'integration_event_id': ie_id}


# ========== APPROVALS & RECONCILIATION ==========

@router.post("/api/adjustments")
async def create_adjustment(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator', 'manager')
    body = await request.json()
    adj_id = str(uuid.uuid4())
    with get_db() as db:
        execute(db,
            "INSERT INTO adjustment_requests (id, inventory_lot_id, outbound_request_line_id, "
            "adjustment_type, qty_before, qty_after, reason_code, notes, status, requested_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (adj_id, body.get('inventory_lot_id'), body.get('outbound_request_line_id'),
             body['adjustment_type'], body.get('qty_before'), body.get('qty_after'),
             body['reason_code'], body.get('notes'), 'pending', user['uid']))
        write_audit(db, user['uid'], 'create', 'adjustment_request', adj_id, None, body)
        db.commit()
    return {'id': adj_id, 'success': True}


@router.put("/api/adjustments/{adj_id}/approve")
async def approve_adjustment_v2(request: Request, adj_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    with get_db() as db:
        adj = dict_from_row(fetchone(db, "SELECT * FROM adjustment_requests WHERE id=%s", (adj_id,)))
        if not adj:
            raise HTTPException(status_code=404, detail="Adjustment not found")
        if adj['status'] != 'pending':
            raise HTTPException(status_code=400, detail="Adjustment is not pending")
        execute(db,
            "UPDATE adjustment_requests SET status='approved', approved_by=%s, approved_at=NOW() WHERE id=%s",
            (user['uid'], adj_id))

        # Apply based on adjustment_type
        if adj['adjustment_type'] == 'write_off' and adj.get('inventory_lot_id'):
            lot = dict_from_row(fetchone(db, "SELECT * FROM inventory_lots WHERE id=%s", (adj['inventory_lot_id'],)))
            if lot:
                qty_before = float(lot['qty_on_hand'])
                mov_id = str(uuid.uuid4())
                ikey = f"adj-{adj_id}-writeoff"
                try:
                    execute(db,
                        "INSERT INTO inventory_movements (id, event_idempotency_key, movement_type, inventory_lot_id, "
                        "tracking_id, qty_delta, qty_before, qty_after, action_by, reason_code) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (mov_id, ikey, 'adjust_down', lot['id'], lot['tracking_id'],
                         -qty_before, qty_before, 0.0, user['uid'], 'write_off'))
                    execute(db, "UPDATE inventory_lots SET qty_on_hand=0, status='consumed', updated_at=NOW() WHERE id=%s",
                            (adj['inventory_lot_id'],))
                except Exception:
                    pass

        elif adj['adjustment_type'] == 'split_roll' and adj.get('inventory_lot_id'):
            lot = dict_from_row(fetchone(db, "SELECT * FROM inventory_lots WHERE id=%s", (adj['inventory_lot_id'],)))
            if lot and adj.get('qty_after') is not None:
                original_qty = float(lot['qty_on_hand'])
                new_qty = float(adj['qty_after'])
                remainder = round(original_qty - new_qty, 4)
                if remainder > 0:
                    # Update original lot
                    execute(db, "UPDATE inventory_lots SET qty_on_hand=%s, updated_at=NOW() WHERE id=%s",
                            (new_qty, lot['id']))
                    # Create new lot with remainder
                    new_lot_id = str(uuid.uuid4())
                    new_tracking = f"TRK-SPLIT-{new_lot_id[:8].upper()}"
                    execute(db,
                        "INSERT INTO inventory_lots (id, item_id, tracking_id, batch_no, shade_code, width_value, "
                        "qty_original, qty_on_hand, qty_reserved, warehouse_id, location_id, status, qty_confidence, created_by) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (new_lot_id, lot['item_id'], new_tracking, lot.get('batch_no'), lot.get('shade_code'),
                         lot.get('width_value'), remainder, remainder, 0.0,
                         lot['warehouse_id'], lot.get('location_id'), 'active', 'measured', user['uid']))
                    # Movements
                    mov1_id = str(uuid.uuid4())
                    execute(db,
                        "INSERT INTO inventory_movements (id, event_idempotency_key, movement_type, inventory_lot_id, "
                        "tracking_id, qty_delta, qty_before, qty_after, action_by, reason_code) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (mov1_id, f"adj-{adj_id}-split-orig", 'adjust_down', lot['id'], lot['tracking_id'],
                         -remainder, original_qty, new_qty, user['uid'], 'split_roll'))
                    mov2_id = str(uuid.uuid4())
                    execute(db,
                        "INSERT INTO inventory_movements (id, event_idempotency_key, movement_type, inventory_lot_id, "
                        "tracking_id, qty_delta, qty_before, qty_after, action_by, reason_code) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (mov2_id, f"adj-{adj_id}-split-new", 'receive', new_lot_id, new_tracking,
                         remainder, 0.0, remainder, user['uid'], 'split_roll'))

        elif adj['adjustment_type'] == 'qty_correction' and adj.get('inventory_lot_id') and adj.get('qty_after') is not None:
            lot = dict_from_row(fetchone(db, "SELECT * FROM inventory_lots WHERE id=%s", (adj['inventory_lot_id'],)))
            if lot:
                qty_before = float(lot['qty_on_hand'])
                qty_after = float(adj['qty_after'])
                delta = qty_after - qty_before
                mtype = 'adjust_up' if delta >= 0 else 'adjust_down'
                mov_id = str(uuid.uuid4())
                ikey = f"adj-{adj_id}-apply"
                try:
                    execute(db,
                        "INSERT INTO inventory_movements (id, event_idempotency_key, movement_type, inventory_lot_id, "
                        "tracking_id, qty_delta, qty_before, qty_after, action_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (mov_id, ikey, mtype, lot['id'], lot['tracking_id'],
                         delta, qty_before, qty_after, user['uid']))
                    execute(db, "UPDATE inventory_lots SET qty_on_hand=%s, updated_at=NOW() WHERE id=%s",
                            (qty_after, adj['inventory_lot_id']))
                except Exception:
                    pass

        write_audit(db, user['uid'], 'approve', 'adjustment_request', adj_id,
                    {'status': 'pending'}, {'status': 'approved'})
        db.commit()
    return {'success': True}


@router.put("/api/adjustments/{adj_id}/reject")
async def reject_adjustment_v2(request: Request, adj_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    with get_db() as db:
        adj = dict_from_row(fetchone(db, "SELECT * FROM adjustment_requests WHERE id=%s", (adj_id,)))
        if not adj:
            raise HTTPException(status_code=404, detail="Adjustment not found")
        execute(db,
            "UPDATE adjustment_requests SET status='rejected', approved_by=%s, approved_at=NOW() WHERE id=%s",
            (user['uid'], adj_id))
        write_audit(db, user['uid'], 'reject', 'adjustment_request', adj_id,
                    {'status': 'pending'}, {'status': 'rejected', 'notes': body.get('notes')})
        db.commit()
    return {'success': True}


@router.post("/api/reconciliation/run")
async def run_reconciliation(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'manager')
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    run_type = body.get('run_type', 'manual')

    with get_db() as db:
        run_id = str(uuid.uuid4())
        execute(db,
            "INSERT INTO reconciliation_runs (id, run_type, status, run_by) VALUES (%s,%s,%s,%s)",
            (run_id, run_type, 'running', user['uid']))
        db.commit()

        findings = []

        neg_lots = rows_to_list(fetchall(db,
            "SELECT * FROM inventory_lots WHERE qty_on_hand < 0 AND status='active'"))
        for lot in neg_lots:
            findings.append(('negative_balance', 'critical',
                f"Lot {lot['tracking_id']} has negative balance: {lot['qty_on_hand']}", 'inventory_lot', lot['id']))

        stuck = rows_to_list(fetchall(db,
            "SELECT * FROM outbound_request_lines WHERE status IN ('needs_approval','in_progress','allocated') "
            "AND updated_at < DATE_SUB(NOW(), INTERVAL 24 HOUR)"))
        for line in stuck:
            findings.append(('stuck_line', 'warning',
                f"Line {line['id']} in '{line['status']}' state for >24h", 'outbound_request_line', line['id']))

        low_lots = rows_to_list(fetchall(db,
            "SELECT * FROM inventory_lots WHERE status='active' AND qty_on_hand < 10 AND qty_on_hand > 0"))
        for lot in low_lots:
            findings.append(('low_remainder', 'warning',
                f"Lot {lot['tracking_id']} has only {lot['qty_on_hand']}m remaining - below 10m threshold",
                'inventory_lot', lot['id']))

        unprinted = rows_to_list(fetchall(db,
            "SELECT ct.id, ct.tracking_id FROM cut_transactions ct "
            "LEFT JOIN tag_labels tl ON ct.id = tl.cut_transaction_id "
            "WHERE tl.id IS NULL OR tl.tag_status = 'generated'"))
        for ct in unprinted:
            findings.append(('missing_tag', 'critical',
                f"Cut transaction {ct['id']} has tag in generated state, not yet printed",
                'cut_transaction', ct['id']))

        pending_ie = rows_to_list(fetchall(db,
            "SELECT * FROM integration_events WHERE status='pending' AND created_at < DATE_SUB(NOW(), INTERVAL 1 HOUR)"))
        for ie in pending_ie:
            findings.append(('unsync_event', 'warning',
                f"Integration event {ie['id']} ({ie['event_type']}) pending for >1h",
                'integration_event', ie['id']))

        for ftype, sev, desc, rtype, rid in findings:
            fid = str(uuid.uuid4())
            execute(db,
                "INSERT INTO reconciliation_findings (id, reconciliation_run_id, finding_type, severity, description, resource_type, resource_id, resolution_status) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                (fid, run_id, ftype, sev, desc, rtype, rid, 'open'))

        execute(db,
            "UPDATE reconciliation_runs SET status='completed', findings_count=%s, completed_at=NOW() WHERE id=%s",
            (len(findings), run_id))
        write_audit(db, user['uid'], 'run_reconciliation', 'reconciliation_run', run_id,
                    None, {'findings_count': len(findings), 'run_type': run_type})
        db.commit()
    return {'run_id': run_id, 'findings_count': len(findings), 'success': True}


# ========== MANUAL MEASURING ==========

@router.post("/api/inventory/measure")
async def measure_inventory(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator')
    body = await request.json()
    tracking_id = body.get('tracking_id')
    measured_qty = float(body.get('measured_qty', 0))
    if not tracking_id:
        raise HTTPException(status_code=400, detail="tracking_id required")

    with get_db() as db:
        lot = dict_from_row(fetchone(db, "SELECT * FROM inventory_lots WHERE tracking_id=%s", (tracking_id,)))
        if not lot:
            raise HTTPException(status_code=404, detail="Lot not found")

        qty_before = float(lot['qty_on_hand'])
        delta = round(measured_qty - qty_before, 4)
        mtype = 'adjust_up' if delta >= 0 else 'adjust_down'

        execute(db, "UPDATE inventory_lots SET qty_on_hand=%s, qty_confidence='measured', updated_at=NOW() WHERE id=%s",
                (measured_qty, lot['id']))

        mov_id = str(uuid.uuid4())
        ikey = f"measure-{lot['id']}-{uuid.uuid4().hex[:8]}"
        execute(db,
            "INSERT INTO inventory_movements (id, event_idempotency_key, movement_type, inventory_lot_id, "
            "tracking_id, qty_delta, qty_before, qty_after, action_by, reason_code) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (mov_id, ikey, mtype, lot['id'], tracking_id,
             delta, qty_before, measured_qty, user['uid'], 'manual_measure'))

        write_audit(db, user['uid'], 'measure', 'inventory_lot', lot['id'],
                    {'qty_on_hand': qty_before, 'qty_confidence': lot['qty_confidence']},
                    {'qty_on_hand': measured_qty, 'qty_confidence': 'measured'})
        db.commit()
    return {'success': True, 'qty_before': qty_before, 'qty_after': measured_qty, 'delta': delta}


# ========== RESERVATIONS ==========

@router.get("/api/reservations")
async def get_reservations(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    with get_db() as db:
        reservations = rows_to_list(fetchall(db, """
            SELECT ir.*, il.tracking_id as lot_tracking, i.name as item_name, u.display_name as created_by_name
            FROM inventory_reservations ir
            LEFT JOIN inventory_lots il ON ir.inventory_lot_id = il.id
            LEFT JOIN items i ON il.item_id = i.id
            LEFT JOIN users u ON ir.created_by = u.id
            ORDER BY ir.created_at DESC
        """))
    return {'reservations': reservations}


@router.post("/api/reservations")
async def create_reservation(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator', 'manager')
    body = await request.json()
    res_id = str(uuid.uuid4())
    with get_db() as db:
        lot = dict_from_row(fetchone(db, "SELECT * FROM inventory_lots WHERE id=%s", (body['inventory_lot_id'],)))
        if not lot:
            raise HTTPException(status_code=404, detail="Inventory lot not found")
        available = float(lot['qty_on_hand']) - float(lot.get('qty_reserved') or 0)
        qty = float(body['qty_reserved'])
        if qty > available:
            raise HTTPException(status_code=400, detail=f"Requested qty {qty} exceeds available {available}")

        execute(db,
            "INSERT INTO inventory_reservations (id, inventory_lot_id, outbound_request_line_id, qty_reserved, reason, status, created_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (res_id, body['inventory_lot_id'], body.get('outbound_request_line_id'),
             qty, body.get('reason'), 'pending_approval', user['uid']))
        write_audit(db, user['uid'], 'create', 'inventory_reservation', res_id, None, body)
        db.commit()
    return {'id': res_id, 'success': True}


@router.post("/api/reservations/{res_id}/approve")
async def approve_reservation(request: Request, res_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    with get_db() as db:
        res = dict_from_row(fetchone(db, "SELECT * FROM inventory_reservations WHERE id=%s", (res_id,)))
        if not res:
            raise HTTPException(status_code=404, detail="Reservation not found")
        if res['status'] != 'pending_approval':
            raise HTTPException(status_code=400, detail="Reservation is not pending approval")

        # Actually reserve qty on lot
        execute(db, "UPDATE inventory_lots SET qty_reserved = qty_reserved + %s, updated_at=NOW() WHERE id=%s",
                (res['qty_reserved'], res['inventory_lot_id']))
        execute(db, "UPDATE inventory_reservations SET status='active', approved_by=%s, approved_at=NOW() WHERE id=%s",
                (user['uid'], res_id))
        write_audit(db, user['uid'], 'approve', 'inventory_reservation', res_id,
                    {'status': 'pending_approval'}, {'status': 'active'})
        db.commit()
    return {'success': True}


@router.post("/api/reservations/{res_id}/reject")
async def reject_reservation(request: Request, res_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    with get_db() as db:
        res = dict_from_row(fetchone(db, "SELECT * FROM inventory_reservations WHERE id=%s", (res_id,)))
        if not res:
            raise HTTPException(status_code=404, detail="Reservation not found")
        execute(db, "UPDATE inventory_reservations SET status='cancelled', approved_by=%s, approved_at=NOW() WHERE id=%s",
                (user['uid'], res_id))
        write_audit(db, user['uid'], 'reject', 'inventory_reservation', res_id,
                    {'status': res['status']}, {'status': 'cancelled'})
        db.commit()
    return {'success': True}


# ========== RETURNS ==========

@router.get("/api/returns")
async def get_returns(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'manager')
    with get_db() as db:
        returns = rows_to_list(fetchall(db, """
            SELECT r.*, i.name as item_name, i.sku, w.name as warehouse_name
            FROM returns r
            LEFT JOIN items i ON r.item_id = i.id
            LEFT JOIN warehouses w ON r.warehouse_id = w.id
            ORDER BY r.created_at DESC
        """))
    return {'returns': returns}


@router.post("/api/returns")
async def create_return(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'warehouse_lead', 'warehouse_operator')
    body = await request.json()
    ret_id = str(uuid.uuid4())
    new_tracking = body.get('new_tracking_id') or f"RTN-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}-{ret_id[:6].upper()}"
    qty = float(body['qty_returned'])

    with get_db() as db:
        execute(db,
            "INSERT INTO returns (id, original_tracking_id, new_tracking_id, item_id, qty_returned, return_reason, "
            "return_type, warehouse_id, location_id, status, received_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (ret_id, body.get('original_tracking_id'), new_tracking, body['item_id'], qty,
             body.get('return_reason'), body.get('return_type', 'other'),
             body.get('warehouse_id'), body.get('location_id'), 'pending', user['uid']))

        # Create new inventory lot for the returned fabric
        lot_id = str(uuid.uuid4())
        warehouse_id = body.get('warehouse_id', 'wh-01')
        execute(db,
            "INSERT INTO inventory_lots (id, item_id, tracking_id, qty_original, qty_on_hand, qty_reserved, "
            "warehouse_id, location_id, status, qty_confidence, created_by) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (lot_id, body['item_id'], new_tracking, qty, qty, 0.0,
             warehouse_id, body.get('location_id'), 'active', 'measured', user['uid']))

        mov_id = str(uuid.uuid4())
        execute(db,
            "INSERT INTO inventory_movements (id, event_idempotency_key, movement_type, inventory_lot_id, "
            "tracking_id, qty_delta, qty_before, qty_after, warehouse_to_id, location_to_id, action_by, reason_code) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (mov_id, f"return-{ret_id}", 'receive', lot_id, new_tracking,
             qty, 0.0, qty, warehouse_id, body.get('location_id'), user['uid'], 'return'))

        write_audit(db, user['uid'], 'create', 'return', ret_id, None,
                    {'new_tracking_id': new_tracking, 'qty_returned': qty})
        db.commit()
    return {'id': ret_id, 'new_tracking_id': new_tracking, 'lot_id': lot_id, 'success': True}


# ========== E-COMMERCE CHANNELS ==========

@router.get("/api/channels")
async def get_channels(request: Request):
    user = require_auth(request)
    with get_db() as db:
        channels = rows_to_list(fetchall(db,
            "SELECT * FROM channel_connections ORDER BY created_at DESC"))
        for ch in channels:
            for field in ('api_key_encrypted', 'api_secret_encrypted', 'access_token_encrypted', 'refresh_token_encrypted'):
                if ch.get(field):
                    ch[field] = '***'
    return {'channels': channels}


@router.post("/api/channels")
async def create_channel(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'manager')
    body = await request.json()
    ch_id = str(uuid.uuid4())
    with get_db() as db:
        execute(db,
            "INSERT INTO channel_connections (id, channel_type, shop_name, api_key_encrypted, api_secret_encrypted, "
            "access_token_encrypted, refresh_token_encrypted, shop_url, region, is_active) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (ch_id, body['channel_type'], body.get('shop_name'),
             body.get('api_key'), body.get('api_secret'), body.get('access_token'), body.get('refresh_token'),
             body.get('shop_url'), body.get('region'), body.get('is_active', 1)))
        write_audit(db, user['uid'], 'create', 'channel_connection', ch_id,
                    None, {k: v for k, v in body.items() if 'key' not in k.lower() and 'secret' not in k.lower() and 'token' not in k.lower()})
        db.commit()
    return {'id': ch_id, 'success': True}


@router.put("/api/channels/{channel_id}")
async def update_channel(request: Request, channel_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'manager')
    body = await request.json()
    with get_db() as db:
        before = dict_from_row(fetchone(db, "SELECT * FROM channel_connections WHERE id=%s", (channel_id,)))
        if not before:
            raise HTTPException(status_code=404, detail="Channel not found")
        fields = {k: v for k, v in body.items() if k in (
            'shop_name', 'api_key_encrypted', 'api_secret_encrypted', 'access_token_encrypted',
            'refresh_token_encrypted', 'shop_url', 'region', 'is_active'
        )}
        if 'api_key' in body: fields['api_key_encrypted'] = body['api_key']
        if 'api_secret' in body: fields['api_secret_encrypted'] = body['api_secret']
        if 'access_token' in body: fields['access_token_encrypted'] = body['access_token']
        if 'refresh_token' in body: fields['refresh_token_encrypted'] = body['refresh_token']
        if fields:
            fields['updated_at'] = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
            set_clause = ", ".join(f"{k}=%s" for k in fields)
            execute(db, f"UPDATE channel_connections SET {set_clause} WHERE id=%s", list(fields.values()) + [channel_id])
        write_audit(db, user['uid'], 'update', 'channel_connection', channel_id,
                    None, {k: '***' if 'key' in k or 'secret' in k or 'token' in k else v for k, v in fields.items()})
        db.commit()
    return {'success': True}


@router.delete("/api/channels/{channel_id}")
async def deactivate_channel(request: Request, channel_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'manager')
    with get_db() as db:
        before = dict_from_row(fetchone(db, "SELECT * FROM channel_connections WHERE id=%s", (channel_id,)))
        if not before:
            raise HTTPException(status_code=404, detail="Channel not found")
        execute(db, "UPDATE channel_connections SET is_active=0, updated_at=NOW() WHERE id=%s", (channel_id,))
        write_audit(db, user['uid'], 'deactivate', 'channel_connection', channel_id,
                    {'is_active': 1}, {'is_active': 0})
        db.commit()
    return {'success': True}


@router.post("/api/channels/{channel_id}/sync_orders")
async def sync_channel_orders(request: Request, channel_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'manager')
    with get_db() as db:
        ch = dict_from_row(fetchone(db, "SELECT * FROM channel_connections WHERE id=%s", (channel_id,)))
        if not ch:
            raise HTTPException(status_code=404, detail="Channel not found")
        if not ch.get('is_active'):
            raise HTTPException(status_code=400, detail="Channel is not active")

        adapter = CHANNEL_ADAPTERS.get(ch['channel_type'], {})
        sync_fn = adapter.get('sync_orders', _stub_sync_orders)
        synced, _ = sync_fn(ch, db, user)

        execute(db, "UPDATE channel_connections SET last_sync_at=NOW(), updated_at=NOW() WHERE id=%s", (channel_id,))
        write_audit(db, user['uid'], 'sync_orders', 'channel_connection', channel_id,
                    None, {'synced_count': synced, 'stub': True})
        db.commit()
    return {'success': True, 'stub': True, 'synced_orders': synced,
            'message': f"Stub sync: {synced} orders simulated for {ch['channel_type']} channel '{ch['shop_name']}'"}


@router.post("/api/channels/{channel_id}/push_inventory")
async def push_channel_inventory(request: Request, channel_id: str):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'manager')
    with get_db() as db:
        ch = dict_from_row(fetchone(db, "SELECT * FROM channel_connections WHERE id=%s", (channel_id,)))
        if not ch:
            raise HTTPException(status_code=404, detail="Channel not found")

        mappings = rows_to_list(fetchall(db,
            "SELECT cpm.*, i.sku, "
            "COALESCE(SUM(il.qty_on_hand - COALESCE(il.qty_reserved,0)),0) as available_qty "
            "FROM channel_product_mappings cpm "
            "JOIN items i ON cpm.nexray_item_id = i.id "
            "LEFT JOIN inventory_lots il ON il.item_id = i.id AND il.status='active' "
            "WHERE cpm.channel_connection_id=%s AND cpm.is_active=1 GROUP BY cpm.id",
            (channel_id,)))

        adapter = CHANNEL_ADAPTERS.get(ch['channel_type'], {})
        push_fn = adapter.get('push_inventory', _stub_push_inventory)
        pushed = push_fn(ch, db, mappings)

        execute(db, "UPDATE channel_connections SET last_sync_at=NOW(), updated_at=NOW() WHERE id=%s", (channel_id,))
        write_audit(db, user['uid'], 'push_inventory', 'channel_connection', channel_id,
                    None, {'pushed_count': len(pushed), 'stub': True})
        db.commit()
    return {'success': True, 'stub': True, 'pushed_items': pushed,
            'message': f"Stub push: {len(pushed)} items simulated for {ch['channel_type']} channel '{ch['shop_name']}'"}


@router.get("/api/channel_mappings")
async def get_channel_mappings(request: Request, channel_id: str = None):
    user = require_auth(request)
    with get_db() as db:
        if channel_id:
            mappings = rows_to_list(fetchall(db,
                "SELECT cpm.*, i.sku, i.name as item_name, cc.channel_type, cc.shop_name "
                "FROM channel_product_mappings cpm "
                "JOIN items i ON cpm.nexray_item_id = i.id "
                "JOIN channel_connections cc ON cpm.channel_connection_id = cc.id "
                "WHERE cpm.channel_connection_id=%s ORDER BY cpm.created_at DESC", (channel_id,)))
        else:
            mappings = rows_to_list(fetchall(db,
                "SELECT cpm.*, i.sku, i.name as item_name, cc.channel_type, cc.shop_name "
                "FROM channel_product_mappings cpm "
                "JOIN items i ON cpm.nexray_item_id = i.id "
                "JOIN channel_connections cc ON cpm.channel_connection_id = cc.id "
                "ORDER BY cpm.created_at DESC"))
    return {'mappings': mappings}


@router.post("/api/channel_mappings")
async def create_channel_mapping(request: Request):
    user = require_auth(request)
    require_role(user, 'system_admin', 'inventory_admin', 'manager')
    body = await request.json()
    cpm_id = str(uuid.uuid4())
    with get_db() as db:
        execute(db,
            "INSERT INTO channel_product_mappings (id, channel_connection_id, channel_product_id, channel_sku, nexray_item_id, is_active) "
            "VALUES (%s,%s,%s,%s,%s,%s)",
            (cpm_id, body['channel_connection_id'], body.get('channel_product_id'), body.get('channel_sku'),
             body['nexray_item_id'], body.get('is_active', 1)))
        write_audit(db, user['uid'], 'create', 'channel_product_mapping', cpm_id, None, body)
        db.commit()
    return {'id': cpm_id, 'success': True}


# ========== CHANNEL ADAPTERS ==========

def _stub_sync_orders(channel, db, user):
    stub_orders = [
        {"channel_order_id": f"STUB-{channel['id'][:4]}-{i:03d}", "channel_status": "paid", "items": []}
        for i in range(1, 4)
    ]
    synced = 0
    for order in stub_orders:
        mapping_id = str(uuid.uuid4())
        try:
            execute(db,
                "INSERT INTO channel_order_mappings (id, channel_connection_id, channel_order_id, channel_status, sync_status, raw_order_json) "
                "VALUES (%s,%s,%s,%s,%s,%s)",
                (mapping_id, channel['id'], order['channel_order_id'], order['channel_status'],
                 'synced', json.dumps(order)))
            synced += 1
        except Exception:
            pass
    return synced, stub_orders


def _stub_push_inventory(channel, db, mappings):
    pushed = []
    for m in mappings:
        pushed.append({
            'channel_sku': m['channel_sku'],
            'nexray_sku': m['sku'],
            'available_qty': round(float(m['available_qty']), 2),
            'push_status': 'stub_success'
        })
    return pushed


CHANNEL_ADAPTERS = {
    'shopify': {'sync_orders': _stub_sync_orders, 'push_inventory': _stub_push_inventory},
    'shopee': {'sync_orders': _stub_sync_orders, 'push_inventory': _stub_push_inventory},
    'lazada': {'sync_orders': _stub_sync_orders, 'push_inventory': _stub_push_inventory},
    'tiktokshop': {'sync_orders': _stub_sync_orders, 'push_inventory': _stub_push_inventory},
}


# ========== WEBHOOK RECEIVER ==========

@router.post("/api/webhooks/{channel_type}")
async def webhook_receiver(request: Request, channel_type: str):
    if channel_type not in ('shopify', 'shopee', 'lazada', 'tiktokshop'):
        return JSONResponse({"detail": "Unknown channel type"}, status_code=400)

    body = await request.json()
    event_id = str(uuid.uuid4())

    with get_db() as db:
        execute(db,
            "INSERT INTO integration_events (id, event_type, event_idempotency_key, "
            "payload_json, status, direction) VALUES (%s,%s,%s,%s,%s,%s)",
            (event_id, f'{channel_type}_webhook', f'wh-{uuid.uuid4().hex[:12]}',
             json.dumps(body), 'pending', 'inbound'))
        db.commit()
    return {'received': True, 'event_id': event_id}
